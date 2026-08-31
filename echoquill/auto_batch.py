"""Pro: Auto-batch — paste a block of lines, one per video:

    URL | Title | folder\\subfolder

For each video in turn this downloads and SAVES the video and the audio,
transcribes and saves the transcript, then runs a saved question SET against
that video's own transcript and saves the answers as a Q&A file. Every output
for a line lands in that line's folder, which is created as deep as you type
(each backslash is another nested level). Title and folder are optional.
"""

import os
import gc
import re
import shutil
import threading
import tkinter as tk
from tkinter import ttk

from . import theme, helptip, ask_ai
from . import prompts as _pr


HELP = (
    "One line per video:\n"
    "    URL | Title | folder\\subfolder\n\n"
    "- Title and folder are optional (URL is all you need).\n"
    "- Blank title = use the video's own title.\n"
    "- Folder: a full path (C:\\...), or just a name to nest under your\n"
    "  Transcriptions folder. Each backslash makes a deeper folder, created\n"
    "  automatically if it doesn't exist.\n"
    "- Blank lines and lines starting with # are ignored.\n\n"
    "For every video the saved video, audio, transcript and Q&A all go into\n"
    "that line's folder. Your chosen question set then runs against the\n"
    "transcript and the answers save as '<name> - Q&A.txt'.\n\n"
    "Everything runs on your PC; each video's answers come only from that\n"
    "video's own transcript."
)


def _segments(raw):
    return [p.strip() for p in re.split(r"[\\/]+", raw.strip()) if p.strip()]


def normalize_name(s):
    """Clean a title/folder-segment so it's a safe, tidy file/folder name.
    Never used on the URL. Colon -> ' -', parentheses dropped, other illegal
    filename characters -> '-'. Whitespace/dashes collapsed."""
    s = (s or "").strip()
    s = s.replace(":", " -")
    s = re.sub(r"[()]", "", s)
    s = re.sub(r'[<>"/\\|?*]', "-", s)
    s = re.sub(r"\s{2,}", " ", s)
    s = re.sub(r"-{2,}", "-", s)
    return s.strip(" -").strip()


def normalize_folder(s):
    """Normalize each level of a (possibly nested) folder path, keep backslashes."""
    segs = [normalize_name(x) for x in _segments(s or "")]
    return "\\".join(x for x in segs if x)


def resolve_folder(cfg, raw, create=True):
    """Return the directory for this line's folder field. Created unless
    create=False (peek at a path without making an empty folder)."""
    from .media_gui import transcripts_dir
    raw = (raw or "").strip().strip('"')
    if not raw:
        return transcripts_dir(cfg)
    m = re.match(r"^([A-Za-z]:[\\/]|\\\\|/)", raw)
    if m:                                   # absolute path
        root = m.group(1)
        rest = [normalize_name(x) for x in _segments(raw[len(root):])]
        rest = [x for x in rest if x]
        d = root + os.sep.join(rest) if rest else root
    else:                                   # relative -> nest under Transcriptions
        rel = [normalize_name(x) for x in _segments(raw)]
        rel = [x for x in rel if x]
        d = os.path.join(transcripts_dir(cfg), *rel) if rel else transcripts_dir(cfg)
    if create:
        os.makedirs(d, exist_ok=True)
    return d


def parse_lines(text):
    """Return [{url, title, folder}, ...] for each usable line."""
    rows = []
    for ln in (text or "").splitlines():
        s = ln.strip()
        if not s or s.startswith("#"):
            continue
        parts = [p.strip() for p in s.split("|")]
        url = parts[0] if parts else ""
        if not url:
            continue
        rows.append({
            "url": url,
            "title": parts[1] if len(parts) > 1 else "",
            "folder": parts[2] if len(parts) > 2 else "",
        })
    return rows


def _dated(name):
    """Prefix a file name with today's date, YYYY-MM-DD-, so files sort by day."""
    import datetime
    return datetime.date.today().strftime("%Y-%m-%d") + "-" + (name or "video")


def _channel_tab_url(channel, tab):
    c = (channel or "").strip().rstrip("/")
    low = c.lower()
    for suff in ("/videos", "/shorts", "/streams", "/live", "/featured"):
        if low.endswith(suff):
            c = c[:-len(suff)]; break
    low = c.lower()
    if c.startswith("http://") or c.startswith("https://"):
        pass
    elif c.startswith("@"):
        c = "https://www.youtube.com/" + c
    elif low.startswith("youtube.com") or low.startswith("www.youtube.com"):
        c = "https://" + c
    else:
        c = "https://www.youtube.com/@" + c
    return c + "/" + tab


def _entry_upload_ts(e, cfg):
    """Best-effort unix timestamp of a search entry's upload; fetches the single
    video's metadata only if the flat entry doesn't already carry a date."""
    import time
    ts = e.get("timestamp") or e.get("release_timestamp")
    if ts:
        try:
            return float(ts)
        except Exception:
            pass
    ud = e.get("upload_date")
    if ud and len(str(ud)) == 8:
        try:
            import datetime
            return datetime.datetime.strptime(str(ud), "%Y%m%d").timestamp()
        except Exception:
            pass
    url = e.get("url") or e.get("webpage_url")
    if not url:
        return None
    try:
        import yt_dlp
        opts = {"quiet": True, "no_warnings": True, "skip_download": True}
        from .media_gui import _apply_proxy
        _apply_proxy(opts, cfg)
        with yt_dlp.YoutubeDL(opts) as ydl:
            info = ydl.extract_info(url, download=False)
        ts = info.get("timestamp") or info.get("release_timestamp")
        if ts:
            return float(ts)
        ud = info.get("upload_date")
        if ud and len(str(ud)) == 8:
            import datetime
            return datetime.datetime.strptime(str(ud), "%Y%m%d").timestamp()
    except Exception:
        return None
    return None


def _yt_sp(upload_days, sort_date=True):
    """Build a YouTube 'sp' search-filter token: optional sort-by-upload-date
    plus the nearest upload-date bucket for the requested window. YouTube then
    filters by date server-side (fast + reliable) - no per-video date lookups.
    Buckets: today / this week / this month / this year (YouTube has no exact
    N-day option, so 60/90-day maps to 'this year')."""
    import base64
    buf = b""
    if sort_date:
        buf += b"\x08\x02"                       # field1 sort = 2 (upload date)
    d = int(upload_days or 0)
    bucket = 0 if d <= 0 else 2 if d <= 1 else 3 if d <= 7 else 4 if d <= 31 else 5
    if bucket:
        sub = bytes([0x08, bucket])               # submsg field1 = date bucket
        buf += bytes([0x12, len(sub)]) + sub
    return base64.urlsafe_b64encode(buf).decode()


def fetch_search_filtered(query, cfg, types=("Video",), duration="Any",
                          upload_days=0, sort="Relevance", n=25,
                          log=lambda s: None):
    """YouTube search with client-side type/duration/upload-window filtering.
    Returns [(url, title), ...]. sort: 'Relevance' or 'Upload date'."""
    import yt_dlp
    date_sorted = str(sort).lower().startswith("upload")
    pull = min(max(int(n) * 3, int(n)), 120)
    opts = {"quiet": True, "no_warnings": True, "extract_flat": True,
            "skip_download": True, "playlistend": pull}
    cf = ((cfg or {}).get("yt_cookies_file", "") or "").strip()
    if cf and os.path.exists(cf):
        opts["cookiefile"] = cf
    from .media_gui import _apply_proxy
    _apply_proxy(opts, cfg)
    from urllib.parse import quote
    sp = _yt_sp(int(upload_days or 0), sort_date=date_sorted)
    target = ("https://www.youtube.com/results?search_query=" + quote(query)
              + ("&sp=" + quote(sp) if sp else ""))
    try:
        with yt_dlp.YoutubeDL(opts) as ydl:
            info = ydl.extract_info(target, download=False)
    except Exception as e:
        log(f"  search error: {str(e)[:120]}")
        return []
    entries = [e for e in (info.get("entries") or []) if e]
    log(f"  YouTube returned {len(entries)} raw result(s); "
        f"filtering by type/duration/upload-window…")
    want = set(t.lower() for t in (types or ["Video"]))
    drop_type = drop_dur = 0
    out = []
    for e in entries:
        if not e:
            continue
        u = e.get("url") or e.get("webpage_url")
        if not u and e.get("id"):
            u = "https://www.youtube.com/watch?v=" + e["id"]
        if not u:
            continue
        t = (e.get("title") or "").strip()
        dur = e.get("duration")
        live = (e.get("live_status") or "")
        is_live = live in ("is_live", "is_upcoming", "post_live") or bool(
            e.get("is_live"))
        is_short = ("/shorts/" in u) or (dur is not None and dur <= 60)
        kind = "live" if is_live else ("shorts" if is_short else "video")
        if kind not in want:
            drop_type += 1
            continue
        if duration and duration != "Any" and dur is not None:
            if duration.startswith("Under") and not dur < 240:
                drop_dur += 1
                continue
            if duration.startswith("4") and not (240 <= dur <= 1200):
                drop_dur += 1
                continue
            if duration.startswith("Over") and not dur > 1200:
                drop_dur += 1
                continue
        out.append((u, t))
        if len(out) >= int(n):
            break
    log(f"  kept {len(out)} · dropped: type {drop_type}, duration {drop_dur} "
        f"(date filtered by YouTube server-side)")
    return out


def fetch_playlist(url, limit, cfg):
    """List a YouTube playlist's videos (metadata only, no download).
    Returns [(url, title), ...]."""
    import yt_dlp
    opts = {"quiet": True, "no_warnings": True, "extract_flat": True,
            "skip_download": True}
    if limit:
        opts["playlistend"] = int(limit)
    cf = ((cfg or {}).get("yt_cookies_file", "") or "").strip()
    if cf and os.path.exists(cf):
        opts["cookiefile"] = cf
    from .media_gui import _apply_proxy
    _apply_proxy(opts, cfg)
    try:
        with yt_dlp.YoutubeDL(opts) as ydl:
            info = ydl.extract_info(url, download=False)
    except Exception:
        return []
    entries = info.get("entries") or []
    flat = []
    for e in entries:
        if e and e.get("entries"):
            flat.extend(e["entries"])
        elif e:
            flat.append(e)
    out = []
    for e in flat:
        if not e:
            continue
        u = e.get("url") or e.get("webpage_url")
        if not u and e.get("id"):
            u = "https://www.youtube.com/watch?v=" + e["id"]
        t = (e.get("title") or "").strip()
        if u:
            out.append((u, t))
    return out


def fetch_channel(channel, kind, limit, cfg):
    """List a YouTube channel's Videos/Shorts/Lives (metadata only, no
    download). Returns [(url, title), ...]."""
    import yt_dlp
    tab = {"Videos": "videos", "Shorts": "shorts",
           "Lives": "streams"}.get(kind, "videos")
    target = _channel_tab_url(channel, tab)
    opts = {"quiet": True, "no_warnings": True, "extract_flat": True,
            "skip_download": True}
    if limit:
        opts["playlistend"] = int(limit)
    cf = ((cfg or {}).get("yt_cookies_file", "") or "").strip()
    if cf and os.path.exists(cf):
        opts["cookiefile"] = cf
    from .media_gui import _apply_proxy
    _apply_proxy(opts, cfg)
    with yt_dlp.YoutubeDL(opts) as ydl:
        info = ydl.extract_info(target, download=False)
    flat = []
    for e in (info.get("entries") or []):
        if e and e.get("entries"):
            flat.extend(e["entries"])
        elif e:
            flat.append(e)
    out = []
    for e in flat:
        if not e:
            continue
        u = e.get("url") or e.get("webpage_url")
        if not u and e.get("id"):
            u = "https://www.youtube.com/watch?v=" + e["id"]
        t = (e.get("title") or "").strip()
        if u:
            out.append((u, t))
    if limit:
        out = out[:int(limit)]
    return out


def fetch_search(query, n, cfg):
    """YouTube search (metadata only). Returns [(url, title), ...]."""
    import yt_dlp
    opts = {"quiet": True, "no_warnings": True, "extract_flat": True,
            "skip_download": True}
    cf = ((cfg or {}).get("yt_cookies_file", "") or "").strip()
    if cf and os.path.exists(cf):
        opts["cookiefile"] = cf
    from .media_gui import _apply_proxy
    _apply_proxy(opts, cfg)
    from urllib.parse import quote
    opts["playlistend"] = int(n)
    target = ("https://www.youtube.com/results?search_query=" + quote(query))
    with yt_dlp.YoutubeDL(opts) as ydl:
        info = ydl.extract_info(target, download=False)
    out = []
    for e in (info.get("entries") or []):
        if not e:
            continue
        u = e.get("url") or e.get("webpage_url")
        if not u and e.get("id"):
            u = "https://www.youtube.com/watch?v=" + e["id"]
        t = (e.get("title") or "").strip()
        if u:
            out.append((u, t))
    return out


class AutoBatchWindow:
    def __init__(self, parent, cfg, research=False):
        self.cfg = cfg
        self._cancel = False
        self._busy = False
        self._eng = None

        self.win = tk.Toplevel(parent)
        self.win.title("EchoQuill — Auto-batch + Ask AI")
        self.win.geometry("740x660")
        theme.apply(self.win)

        top = ttk.Frame(self.win)
        top.pack(fill="x", padx=18, pady=(14, 2))
        ttk.Label(top, text="Auto-batch — transcribe + Ask AI",
                  style="Title.TLabel").pack(side="left")
        helptip.attach(self.win, top, "Auto-batch - help", HELP).pack(
            side="left", padx=8)

        nb = ttk.Notebook(self.win)
        nb.pack(fill="both", expand=True, padx=12, pady=(6, 4))
        self._nb = nb
        tab_build = ttk.Frame(nb)
        tab_run = ttk.Frame(nb)
        self._tab_run = tab_run
        nb.add(tab_build, text="  Build list  ")
        nb.add(tab_run, text="  Run options  ")

        # ---------- Build list tab ----------
        brow = ttk.Frame(tab_build); brow.pack(fill="x", padx=8, pady=(8, 2))
        ttk.Button(brow, text="▦ Build from columns…",
                   command=self._open_grid).pack(side="left")
        ttk.Button(brow, text="⭱ Load .xlsx…",
                   command=self._load_xlsx).pack(side="left", padx=6)
        ttk.Label(tab_build, style="Dim.TLabel", wraplength=720, text=(
            "One line per video:   URL | Title | folder\\subfolder      "
            "(Title and folder optional; blank title uses the video's own "
            "title).")).pack(anchor="w", padx=8)
        self.box = theme.dark_text(tab_build, wrap="none", height=11)
        self.box.pack(fill="both", expand=True, padx=8, pady=(6, 2))
        self.box.bind("<KeyRelease>", lambda e: self._recount())
        helptip.tip(self.box, "Paste your list here, one video per line: "
                    "URL | Title | folder. See the ? above for the full format.")
        self.count = ttk.Label(tab_build, text="0 videos", style="Dim.TLabel")
        self.count.pack(anchor="w", padx=8, pady=(0, 6))

        # ---------- Run options tab ----------
        opt = ttk.Frame(tab_run)
        opt.pack(fill="x", padx=8, pady=(10, 2))
        ttk.Label(opt, text="Question set:").pack(side="left")
        self.setvar = tk.StringVar(value="—")
        self.setmenu = ttk.OptionMenu(opt, self.setvar, "—")
        self.setmenu.configure(width=24)
        self.setmenu.pack(side="left", padx=(6, 8))
        self._refresh_sets()
        ttk.Button(opt, text="⚙ Manage sets…",
                   command=self._manage_sets).pack(side="left", padx=(0, 12))
        helptip.tip(self.setmenu, "The saved set of questions to ask every "
                    "video. Create sets in Ask AI → 'Ask several' → 'Save "
                    "checked as set'.")
        trow = ttk.Frame(tab_run)
        trow.pack(fill="x", padx=8, pady=(8, 0))
        ttk.Label(trow, text="Transcript source:").pack(side="left")
        self.transcript_mode = tk.StringVar(value="Whisper (accurate)")
        _tm = ttk.OptionMenu(trow, self.transcript_mode, "Whisper (accurate)",
                             "Whisper (accurate)", "YouTube captions (fast)")
        _tm.pack(side="left", padx=(6, 0))
        helptip.tip(_tm, "Whisper: our local speech-to-text - most accurate, "
                    "works on any site. YouTube captions: pulls the video's "
                    "existing captions instead (fast, no download); falls back "
                    "to Whisper if a video has none.")

        self.save_video = tk.BooleanVar(value=True)
        self.save_audio = tk.BooleanVar(value=True)
        self.save_desc = tk.BooleanVar(value=False)
        saverow = ttk.Frame(tab_run)
        saverow.pack(fill="x", padx=8, pady=(8, 2))
        ttk.Checkbutton(saverow, text="Save video",
                        variable=self.save_video).pack(side="left")
        ttk.Checkbutton(saverow, text="Save audio",
                        variable=self.save_audio).pack(side="left", padx=(12, 0))
        _cbd = ttk.Checkbutton(saverow, text="Video Description",
                               variable=self.save_desc)
        _cbd.pack(side="left", padx=(12, 0))
        self.save_comments = tk.BooleanVar(value=False)
        _cbc = ttk.Checkbutton(saverow, text="Comments",
                               variable=self.save_comments)
        _cbc.pack(side="left", padx=(12, 0))
        helptip.tip(_cbc, "Also save the video's comments to its own "
                    "'<name> - Comments.txt' (off by default; it's slow, so it "
                    "grabs the top ~200).")
        helptip.tip(_cbd, "Also save each video's description as "
                    "'<name> - Description.txt' (off by default). Says "
                    "'No video description' if there is none.")

        thr = ttk.Frame(tab_run)
        thr.pack(fill="x", padx=8, pady=(8, 2))
        ttk.Label(thr, text="Pause every").pack(side="left")
        self.thr_n = tk.StringVar(value="5")
        tk.Entry(thr, textvariable=self.thr_n, width=4, bg=theme.FIELD,
                 fg=theme.FG, insertbackground=theme.FG, relief="solid",
                 borderwidth=1).pack(side="left", padx=4)
        ttk.Label(thr, text="videos for").pack(side="left")
        self.thr_s = tk.StringVar(value="60")
        tk.Entry(thr, textvariable=self.thr_s, width=5, bg=theme.FIELD,
                 fg=theme.FG, insertbackground=theme.FG, relief="solid",
                 borderwidth=1).pack(side="left", padx=4)
        ttk.Label(thr, text="seconds   (0 = no pause; memory is released "
                  "during each pause)", style="Dim.TLabel").pack(side="left")

        # ---------- constant: live monitor ----------
        mon = tk.Frame(self.win, bg="#141414", bd=1, relief="solid")
        mon.pack(fill="x", padx=18, pady=(2, 2))
        self.monitor = tk.Label(mon, bg="#141414", fg="#8a8a8a",
                                font=("Segoe UI", 12, "bold"), anchor="w",
                                justify="left", wraplength=700,
                                text="●  Idle — nothing running")
        self.monitor.pack(fill="x", padx=12, pady=8)
        mon.bind("<Configure>", lambda e: self.monitor.configure(
            wraplength=max(220, e.width - 28)))

        # ---------- constant: run controls ----------
        bar = ttk.Frame(self.win)
        bar.pack(fill="x", padx=18, pady=(2, 2))
        self.start_btn = ttk.Button(bar, text="Start", style="Accent.TButton",
                                    command=self._start)
        self.start_btn.pack(side="left")
        self.stop_btn = ttk.Button(bar, text="Stop", command=self._stop,
                                   state="disabled")
        self.stop_btn.pack(side="left", padx=8)
        ttk.Button(bar, text="Clear all",
                   command=self._clear_all).pack(side="left", padx=8)
        ttk.Button(bar, text="Open folder",
                   command=self._open_folder).pack(side="left", padx=8)
        ttk.Button(bar, text="Close", command=self.win.destroy).pack(side="right")

        self.status = ttk.Label(self.win, text="", style="Dim.TLabel")
        self.status.pack(anchor="w", padx=18)
        ttk.Label(self.win, text="Progress", style="Section.TLabel").pack(
            anchor="w", padx=18, pady=(4, 0))
        self.log = theme.dark_text(self.win, wrap="word", height=8)
        self.log.pack(fill="x", padx=18, pady=(2, 12))
        self.win.protocol("WM_DELETE_WINDOW", self._on_close)
        self.win.transient(parent)
        theme.bring_to_front(self.win)

    # ---------- helpers ----------
    def _open_grid(self):
        AutoBatchGrid(self.win, self.cfg, self._apply_grid)

    def _manage_sets(self):
        from . import prompts as _pr
        _pr.manage_sets(self.win, self.cfg, self._refresh_sets)

    def _load_xlsx(self):
        from tkinter import filedialog
        from .media_gui import transcripts_dir
        path = filedialog.askopenfilename(
            parent=self.win, title="Open a saved video list (.xlsx)",
            initialdir=transcripts_dir(self.cfg),
            filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        try:
            data = load_xlsx(path)
        except Exception as e:
            self._set(f"Couldn't open: {e}")
            return
        self.box.delete("1.0", "end")
        self.box.insert("1.0",
                        "\n".join(_to_line(u, t, f) for (u, t, f) in data))
        self._recount()
        self._set(f"Loaded {len(data)} videos from {os.path.basename(path)}.")

    def _apply_grid(self, text, run=False):
        self.box.delete("1.0", "end")
        self.box.insert("1.0", text)
        self._recount()
        try:
            self.win.deiconify()
            self.win.lift()
            self.win.focus_force()
        except Exception:
            pass
        if run:
            self._start()

    def _on_close(self):
        self._cancel = True
        try:
            self._eng = None
            gc.collect()
        except Exception:
            pass
        self.win.destroy()

    def _refresh_sets(self):
        m = self.setmenu["menu"]
        m.delete(0, "end")
        m.add_command(label="—", command=lambda: self.setvar.set("—"))
        for n in _pr.set_names(self.cfg):
            m.add_command(label=n, command=lambda n=n: self.setvar.set(n))

    def _recount(self):
        n = len(parse_lines(self.box.get("1.0", "end")))
        self.count.configure(text=f"{n} video{'s' if n != 1 else ''}")

    def _open_folder(self):
        from .media_gui import transcripts_dir
        try:
            os.startfile(transcripts_dir(self.cfg))
        except Exception as e:
            self._set(str(e))

    def _clear_all(self):
        self.box.delete("1.0", "end")          # the one-line-per-video list
        self.log.delete("1.0", "end")          # the progress log
        self.setvar.set("—")                   # question set back to none
        self.status.configure(text="")
        self._recount()

    def _set(self, msg):
        def _do():
            self.status.configure(text=msg)
            self._set_monitor(msg)
        try:
            self.win.after(0, _do)
        except Exception:
            pass

    def _set_monitor(self, msg):
        low = (msg or "").lower()
        if not msg:
            col, txt = "#8a8a8a", "●  Idle — nothing running"
        elif low.startswith("done") or "processed" in low:
            col, txt = "#30d158", "✓  " + msg
        elif low.startswith("stop"):
            col, txt = "#e0a030", "■  " + msg
        elif "video" in low:
            col, txt = "#4da3ff", "▶  " + msg
        else:
            col, txt = "#4da3ff", msg
        try:
            self.monitor.configure(text=txt, fg=col)
        except Exception:
            pass

    def _log(self, msg):
        def _do():
            self.log.insert("end", msg + "\n")
            self.log.see("end")
        try:
            self.win.after(0, _do)
        except Exception:
            pass

    def _stop(self):
        self._cancel = True
        self._set("Stopping after this step…")

    # ---------- run ----------
    def _start(self):
        if self._busy:
            return
        rows = parse_lines(self.box.get("1.0", "end"))
        if not rows:
            self._set("Paste at least one URL line first.")
            return
        from . import license as _lic
        if not _lic.is_pro(self.cfg):
            self._set("Auto-batch is a Pro feature.")
            return
        name = self.setvar.get()
        questions = _pr.get_set(self.cfg, name) if name not in ("—", "") else []
        if not questions:
            self._set("Pick a question set first (make one in Ask AI → "
                      "'Ask several' → 'Save checked as set').")
            return
        self._cancel = False
        self._busy = True
        self.start_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        threading.Thread(target=self._worker, args=(rows, questions),
                         daemon=True).start()

    def _throttle_pause(self, done_count):
        import time
        try:
            n = int((self.thr_n.get() or "0").strip() or 0)
            sec = int((self.thr_s.get() or "0").strip() or 0)
        except Exception:
            return
        if n <= 0 or sec <= 0 or (done_count % n) != 0:
            return
        self._eng = None          # release the Whisper model during the wait
        gc.collect()
        for rem in range(sec, 0, -1):
            if self._cancel:
                break
            self._set(f"Cooling down {rem}s\u2026 (memory released)")
            time.sleep(1)

    def _worker(self, rows, questions):
        from .media_gui import (download_video, fetch_audio_info, safe_filename,
                                _safe_stem, _unique_path, _keep_awake,
                                _video_description, _video_comments,
                                fetch_captions)
        from .transcriber import Transcriber
        from . import proxy
        total = len(rows)
        done = 0
        if self._eng is None:
            self._eng = Transcriber(self.cfg.get("model", "base"))
        _keep_awake(True)
        try:
            for i, row in enumerate(rows, 1):
                if self._cancel:
                    break
                url, ttl, fld = row["url"], row["title"], row["folder"]
                tmpdir = None
                try:
                    dest = resolve_folder(self.cfg, fld)
                    self._log(f"[{i}/{total}] {url}")
                    self._log(f"    folder: {dest}")

                    if self.cfg.get("di_enabled"):
                        self._set(f"Video {i}/{total}: verifying proxy IP…")
                        tries = int(self.cfg.get("di_verify_tries", 3) or 3)
                        sid, ip = proxy.acquire_verified(
                            self.cfg, tries=tries, log=self._log)
                        if not sid:
                            self._log(f"    proxy: no live IP after {tries} "
                                      "tries — skipping this video")
                            continue

                    # 1) optional fast path: YouTube captions
                    use_captions = self.transcript_mode.get().lower().startswith(
                        "youtube")
                    cap_segs = cap_text = None
                    cap_title = ""
                    if use_captions:
                        self._set(f"Video {i}/{total}: fetching captions…")
                        cap_segs, cap_text, cap_title = fetch_captions(
                            url, self.cfg)
                        if not cap_text:
                            self._log("    no captions found — using Whisper")
                    got_captions = bool(cap_text)

                    # 2) audio: needed for Whisper, or if you want to save audio
                    apath, atitle, tmpdir = None, "", None
                    if (not got_captions) or self.save_audio.get():
                        self._set(f"Video {i}/{total}: downloading audio…")
                        apath, atitle = fetch_audio_info(url, self._set, self.cfg)
                        tmpdir = os.path.dirname(apath)
                    name = normalize_name(
                        ttl or cap_title or atitle or "video") or "video"
                    dname = _dated(name)

                    # 3) full video file (dated)
                    if self.save_video.get():
                        self._set(f"Video {i}/{total}: downloading video…")
                        try:
                            download_video(url, self.cfg, dest, self._set,
                                           name=dname)
                            self._log("    video saved ✓")
                        except Exception as e:
                            self._log(f"    video save failed: {e}")
                            try:
                                from .media_gui import cleanup_parts
                                cleanup_parts(dest, dname)
                            except Exception:
                                pass

                    # 4) audio file (dated)
                    if self.save_audio.get() and apath and os.path.exists(apath):
                        try:
                            ext = os.path.splitext(apath)[1] or ".m4a"
                            adst = _unique_path(os.path.join(
                                dest, _safe_stem(dname) + ext))
                            shutil.copy2(apath, adst)
                            self._log(f"    audio saved: {os.path.basename(adst)} ✓")
                        except Exception as e:
                            self._log(f"    audio save failed: {e}")

                    # 5) transcript: captions or Whisper
                    if self._cancel:
                        break
                    if got_captions:
                        segs, text = cap_segs, cap_text
                        self._log("    transcript from YouTube captions ✓")
                    else:
                        self._set(f"Video {i}/{total}: transcribing…")
                        if self._eng is None:
                            self._eng = Transcriber(self.cfg.get("model", "base"))
                        model = self._eng.load()
                        lang = self.cfg.get("language", "auto")
                        lang = None if lang in ("", "auto") else lang
                        segs, parts = [], []
                        with self._eng._lock:
                            segments, _info = model.transcribe(
                                apath, language=lang, vad_filter=True)
                            for seg in segments:
                                if self._cancel:
                                    break
                                t = seg.text.strip()
                                segs.append((seg.start, t))
                                parts.append(t)
                        text = " ".join(parts).strip()
                    body = text
                    if self.save_desc.get():
                        desc = _video_description(url, self.cfg)
                        body = ("========== VIDEO DESCRIPTION ==========\n"
                                + desc +
                                "\n\n========== TRANSCRIPTION ==========\n"
                                + text)
                    tpath = _unique_path(os.path.join(dest, safe_filename(dname)))
                    with open(tpath, "w", encoding="utf-8") as f:
                        f.write(f"{name}\n{url}\n\n{body}")
                    self._log(f"    transcript saved: {os.path.basename(tpath)} ✓")
                    if self.save_comments.get():
                        try:
                            self._set(f"Video {i}/{total}: fetching comments…")
                            cmts = _video_comments(url, self.cfg)
                            cpath = _unique_path(os.path.join(
                                dest, safe_filename(dname + " - Comments")))
                            with open(cpath, "w", encoding="utf-8") as fc:
                                fc.write(f"{name}\n{url}\n\n{cmts}")
                            self._log("    comments saved ✓")
                        except Exception as e:
                            self._log(f"    comments failed: {e}")

                    # 5) run the question set, grounded in THIS transcript
                    if self._cancel:
                        break
                    qa = []
                    for qi, q in enumerate(questions, 1):
                        if self._cancel:
                            break
                        self._set(f"Video {i}/{total}: question "
                                  f"{qi}/{len(questions)}…")
                        ans = ask_ai.ask(q, segs, self.cfg, title=name, url=url)
                        qa.append(f"{'*' * 50}\n{'*' * 50}\nQ: {q}\n\nA: {ans}\n")
                    if qa:
                        qpath = _unique_path(os.path.join(
                            dest, safe_filename(dname + " - Q&A")))
                        with open(qpath, "w", encoding="utf-8") as f:
                            f.write(f"{name}\n{url}\n\n" + "\n".join(qa))
                        self._log(f"    Q&A saved: {os.path.basename(qpath)} ✓")
                    done += 1
                    self._log(f"    [{i}/{total}] done ✓")
                except Exception as e:
                    self._log(f"    ERROR on this video: {e}")
                finally:
                    try:
                        proxy.clear_active_port()
                    except Exception:
                        pass
                    if tmpdir:
                        shutil.rmtree(tmpdir, ignore_errors=True)
                    segs = parts = text = None   # drop this video's data
                    gc.collect()                 # quick between-video cleanup
                if not self._cancel and i < total:
                    self._throttle_pause(i)      # release memory + pause
        finally:
            _keep_awake(False)
            self._eng = None      # unload the Whisper model when the batch ends
            gc.collect()
        msg = (f"Stopped — {done}/{total} finished."
               if self._cancel else
               f"Done ✓ — {done}/{total} videos processed.")
        self._set(msg)
        self._log(msg)
        self._busy = False
        try:
            self.win.after(0, lambda: (self.start_btn.configure(state="normal"),
                                       self.stop_btn.configure(state="disabled")))
        except Exception:
            pass


# ============================================================================
# Column builder — a bigger pop-up with a 3-column editable grid that saves to
# a real .xlsx, then loads the rows straight into the Auto-batch box.
# ============================================================================

GRID_HELP = (
    "Paste a whole column at a time: all your URLs in the first box, all the "
    "titles in the second, all the folders in the third - one per line. Line 1 "
    "of each box lines up as video 1, line 2 as video 2, and so on. The counts "
    "under the boxes help you keep the three columns even.\n\n"
    "Title and Folder are optional: blank title = the video's own title; a "
    "folder nests under Transcriptions, each backslash a deeper level.\n\n"
    "Save writes a real .xlsx backup (pick a folder, type a name - no extension "
    "needed). Start loads the list into Auto-batch. Load .xlsx reopens a saved "
    "list. The small clear buttons wipe one column; Clear all wipes everything."
)


def _to_line(u, t, f):
    """Build a 'URL | Title | Folder' line, trimming only trailing blanks."""
    parts = [u, t, f]
    while len(parts) > 1 and not parts[-1]:
        parts.pop()
    return " | ".join(parts)


def _grid_dir(cfg):
    from .media_gui import transcripts_dir
    return transcripts_dir(cfg)


def save_xlsx(path, rows):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "videos"
    ws.append(["URL", "Title", "Folder"])
    for u, t, f in rows:
        ws.append([u, t, f])
    wb.save(path)


def load_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = [("" if c is None else str(c)).strip() for c in row]
        vals = (vals + ["", "", ""])[:3]
        if i == 0 and vals[0].lower() in ("url", "link"):
            continue
        if vals[0]:
            out.append((vals[0], vals[1], vals[2]))
    return out


class AutoBatchGrid:
    """Three side-by-side paste columns (URL / Title / Folder), each with its
    own vertical + horizontal scrollbar. Paste a whole column at a time; line N
    of each box is video N. Save writes a real .xlsx; Start loads the list into
    Auto-batch (only after a save)."""

    def __init__(self, parent, cfg, on_apply):
        self.cfg = cfg
        self.on_apply = on_apply
        self._saved = False
        self._saved_path = ""

        self.win = tk.Toplevel(parent)
        self.win.title("EchoQuill — Build video list (columns)")
        self.win.geometry("740x660")
        theme.apply(self.win)

        top = ttk.Frame(self.win)
        top.pack(fill="x", padx=16, pady=(12, 2))
        ttk.Label(top, text="Build your list — paste a column at a time",
                  style="Title.TLabel").pack(side="left")
        helptip.attach(self.win, top, "Columns - help", GRID_HELP).pack(
            side="left", padx=8)
        ttk.Button(top, text="Clear all",
                   command=self._clear_all).pack(side="right")

        chrow = ttk.Frame(self.win)
        chrow.pack(fill="x", padx=16, pady=(2, 1))
        ttk.Label(chrow, text="Add from:").pack(side="left")
        self.source_var = tk.StringVar(value="YouTube channel")
        ttk.OptionMenu(chrow, self.source_var, "YouTube channel",
                       "YouTube channel", "Search YouTube").pack(
                       side="left", padx=(6, 6))
        self.chan_var = tk.StringVar()
        tk.Entry(chrow, textvariable=self.chan_var, bg=theme.FIELD, fg=theme.FG,
                 insertbackground=theme.FG, relief="solid", borderwidth=1).pack(
                 side="left", fill="x", expand=True, padx=(0, 6))
        ttk.Button(chrow, text="Fetch", style="Accent.TButton",
                   command=self._fetch_channel).pack(side="left")

        chrow2 = ttk.Frame(self.win)
        chrow2.pack(fill="x", padx=16, pady=(0, 2))
        ttk.Label(chrow2, text="Type:").pack(side="left")
        self.kind_var = tk.StringVar(value="Videos")
        ttk.OptionMenu(chrow2, self.kind_var, "Videos", "Videos", "Shorts",
                       "Lives", "All").pack(side="left", padx=(4, 12))
        ttk.Label(chrow2, text="Keyword:").pack(side="left")
        self.keyword_var = tk.StringVar()
        tk.Entry(chrow2, textvariable=self.keyword_var, width=16, bg=theme.FIELD,
                 fg=theme.FG, insertbackground=theme.FG, relief="solid",
                 borderwidth=1).pack(side="left", padx=(4, 12))
        ttk.Label(chrow2, text="How many:").pack(side="left")
        self.count_var = tk.StringVar(value="10")
        ttk.OptionMenu(chrow2, self.count_var, "10", "5", "10", "15", "All",
                       "Other\u2026", command=self._count_pick).pack(
                       side="left", padx=(4, 10))
        self.chan_status = ttk.Label(chrow2, style="Dim.TLabel", text="")
        self.chan_status.pack(side="left", padx=(6, 0))

        cols = ttk.Frame(self.win)
        cols.pack(fill="both", expand=True, padx=12, pady=(4, 2))
        self.url_txt = self._column(cols, "URL")
        self.title_txt = self._column(cols, "Title")
        self.folder_txt = self._column(cols, "Folder")

        self.count = ttk.Label(self.win, style="Dim.TLabel",
                               text="URLs: 0 · Titles: 0 · Folders: 0")
        self.count.pack(anchor="w", padx=16, pady=(2, 2))

        bar = ttk.Frame(self.win)
        bar.pack(fill="x", padx=16, pady=(2, 8))
        ttk.Button(bar, text="Load .xlsx…", command=self._load).pack(side="left")
        ttk.Button(bar, text="Folder \u2192 all rows",
                   command=self._apply_folder).pack(side="left", padx=8)
        ttk.Button(bar, text="Start ▸", style="Accent.TButton",
                   command=self._start_run).pack(side="right")
        ttk.Button(bar, text="Save", command=self._save).pack(side="right", padx=8)
        ttk.Button(bar, text="Close",
                   command=self.win.destroy).pack(side="right", padx=(0, 8))
        self.status = ttk.Label(self.win, style="Dim.TLabel", text="")
        self.status.pack(anchor="w", padx=16, pady=(0, 8))
        self.win.transient(parent)
        theme.bring_to_front(self.win)

    def _apply_folder(self):
        """Take the first non-empty folder and put it on every URL row."""
        folder = ""
        for ln in self.folder_txt.get("1.0", "end").splitlines():
            if ln.strip():
                folder = ln.strip(); break
        urls = [u for u in self.url_txt.get("1.0", "end").splitlines()
                if u.strip()]
        if not urls:
            self.status.configure(text="Add URLs first."); return
        if not folder:
            self.status.configure(
                text="Put a folder in the Folder column's first row first.")
            return
        self.folder_txt.delete("1.0", "end")
        self.folder_txt.insert("1.0", "\n".join([folder] * len(urls)))
        self._recount()
        self.status.configure(
            text=f"Applied '{folder}' to all {len(urls)} rows.")

    def _count_pick(self, v):
        if v == "Other\u2026":
            from tkinter import simpledialog
            n = simpledialog.askstring("How many",
                                       "How many videos?", parent=self.win)
            self.count_var.set(n.strip() if (n and n.strip().isdigit())
                               else "10")

    def _fetch_channel(self):
        import threading
        q = self.chan_var.get().strip()
        is_search = self.source_var.get().lower().startswith("search")
        if not q:
            self.chan_status.configure(
                text=("Type a search term first." if is_search
                      else "Paste a channel URL or @handle first."))
            return
        kind = self.kind_var.get()
        keyword = self.keyword_var.get().strip().lower()
        cv = self.count_var.get().strip().lower()
        want = None if cv == "all" else (int(cv) if cv.isdigit() else 10)
        self.chan_status.configure(
            text=("Searching\u2026" if is_search else f"Fetching {kind}\u2026"))

        def run():
            try:
                if is_search:
                    items = fetch_search(q, want or 25, self.cfg)
                else:
                    fetch_n = (min(want * 8, 300) if (want and keyword)
                               else want)
                    if kind == "All":
                        items, seen = [], set()
                        for _k in ("Videos", "Shorts", "Lives"):
                            for (u, t) in fetch_channel(q, _k, fetch_n,
                                                        self.cfg):
                                if u not in seen:
                                    seen.add(u); items.append((u, t))
                    else:
                        items = fetch_channel(q, kind, fetch_n, self.cfg)
                    if keyword:
                        items = [(u, t) for (u, t) in items
                                 if keyword in (t or "").lower()]
                    if want:
                        items = items[:want]
            except Exception as e:
                self.win.after(0, lambda e=e: self.chan_status.configure(
                    text=f"Fetch failed: {str(e)[:70]}"))
                return

            def fill():
                if not items:
                    self.chan_status.configure(
                        text="No matches." if keyword else "Nothing found.")
                    return
                us = "\n".join(i[0] for i in items)
                ts = "\n".join(i[1] for i in items)
                if self.url_txt.get("1.0", "end").strip():
                    self.url_txt.insert("end", "\n" + us)
                    self.title_txt.insert("end", "\n" + ts)
                else:
                    self.url_txt.insert("1.0", us)
                    self.title_txt.insert("1.0", ts)
                self._recount()
                label = "search results" if is_search else kind
                self.chan_status.configure(text=f"Added {len(items)} {label}.")
            self.win.after(0, fill)
        threading.Thread(target=run, daemon=True).start()

    def _column(self, parent, name):
        col = ttk.Frame(parent)
        col.pack(side="left", fill="both", expand=True, padx=4)
        hdr = ttk.Frame(col)
        hdr.pack(fill="x")
        ttk.Label(hdr, text=name, style="Section.TLabel").pack(side="left")
        ttk.Button(hdr, text="✕", width=2,
                   command=lambda: self._clear_one(name)).pack(side="right")
        wrap = ttk.Frame(col)
        wrap.pack(fill="both", expand=True)
        txt = tk.Text(wrap, wrap="none", width=10, height=18, bg=theme.FIELD,
                      fg=theme.FG, insertbackground=theme.FG, relief="solid",
                      borderwidth=1, font=("Segoe UI", 10))
        vs = ttk.Scrollbar(wrap, orient="vertical", command=txt.yview)
        hs = ttk.Scrollbar(wrap, orient="horizontal", command=txt.xview)
        txt.configure(yscrollcommand=vs.set, xscrollcommand=hs.set)
        txt.grid(row=0, column=0, sticky="nsew")
        vs.grid(row=0, column=1, sticky="ns")
        hs.grid(row=1, column=0, sticky="ew")
        wrap.rowconfigure(0, weight=1)
        wrap.columnconfigure(0, weight=1)
        txt.edit_modified(False)
        txt.bind("<<Modified>>", lambda e, w=txt: self._on_modified(w))
        return txt

    def _on_modified(self, t):
        if t.edit_modified():
            t.edit_modified(False)      # reset so it fires again next change
            self._recount()

    def _boxes(self):
        return {"URL": self.url_txt, "Title": self.title_txt,
                "Folder": self.folder_txt}

    def _lines(self, txt):
        return txt.get("1.0", "end").splitlines()

    def _clear_one(self, name):
        self._boxes()[name].delete("1.0", "end")
        self._recount()

    def _clear_all(self):
        for t in self._boxes().values():
            t.delete("1.0", "end")
        self._recount()
        self.status.configure(text="Cleared.")

    def _recount(self):
        nu = len([x for x in self._lines(self.url_txt) if x.strip()])
        nt = len([x for x in self._lines(self.title_txt) if x.strip()])
        nf = len([x for x in self._lines(self.folder_txt) if x.strip()])
        warn = ""
        if nt and nt != nu:
            warn = "  ⚠ titles ≠ URLs"
        elif nf and nf != nu:
            warn = "  ⚠ folders ≠ URLs"
        self.count.configure(
            text=f"URLs: {nu} · Titles: {nt} · Folders: {nf}{warn}")
        self._saved = False

    def _rows(self):
        us = self._lines(self.url_txt)
        ts = self._lines(self.title_txt)
        fs = self._lines(self.folder_txt)
        n = max(len(us), len(ts), len(fs))
        out = []
        for i in range(n):
            u = us[i].strip() if i < len(us) else ""
            if not u:
                continue
            t = ts[i].strip() if i < len(ts) else ""
            f = fs[i].strip() if i < len(fs) else ""
            out.append((u, normalize_name(t), normalize_folder(f)))
        return out

    def _fill_cols(self, data):
        for t in self._boxes().values():
            t.delete("1.0", "end")
        self.url_txt.insert("1.0", "\n".join(d[0] for d in data))
        self.title_txt.insert("1.0", "\n".join(d[1] for d in data))
        self.folder_txt.insert("1.0", "\n".join(d[2] for d in data))
        self._recount()

    def _load(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            parent=self.win, title="Open a saved video list",
            initialdir=_grid_dir(self.cfg),
            filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        try:
            data = load_xlsx(path)
        except Exception as e:
            self.status.configure(text=f"Couldn't open: {e}")
            return
        self._fill_cols(data)
        self.status.configure(
            text=f"Loaded {len(data)} rows from {os.path.basename(path)}.")

    def _save(self):
        from tkinter import filedialog
        data = self._rows()
        if not data:
            self.status.configure(text="Add at least one URL first.")
            return
        path = filedialog.asksaveasfilename(
            parent=self.win, title="Save this list — pick a folder and name",
            initialdir=_grid_dir(self.cfg), defaultextension=".xlsx",
            initialfile="video-list", filetypes=[("Excel", "*.xlsx")])
        if not path:
            self.status.configure(text="Save cancelled.")
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            save_xlsx(path, data)
        except Exception as e:
            self.status.configure(text=f"Save failed: {e}")
            return
        self._fill_cols(data)           # show the cleaned/normalized values
        self._saved = True
        self._saved_path = path
        self.status.configure(
            text=f"Saved ✓  {len(data)} rows → {os.path.basename(path)}. "
                 "Now press Start.")

    def _start_run(self):
        data = self._rows()
        if not data:
            self.status.configure(text="Add at least one URL first.")
            return
        self.on_apply("\n".join(_to_line(u, t, f) for (u, t, f) in data),
                      run=True)
        self.win.destroy()
